#!/usr/bin/env python
# -*- coding: utf-8 -*-
'''
Calcule la balance analytique a partir des brouillards de banque
USAGE     : $python brouillards2balance.py
SETUP     : brouillards2balance.cfg
OUTPUT    : banlance_new.xlsx
Copyright : Croix-Rouge Francaise 2016 (French Red Cross)
Author    : Julien Borghetti
Date      : 2016/10/02
'''

import openpyxl
from openpyxl.styles import Font
import subprocess
import re
import calendar
import datetime
import ConfigParser
import logging
from logging.handlers import RotatingFileHandler

# import ipdb
# import sys

antennes = [3969, 4010, 4011, 4012, 4013, 4015, 4016]
ressources = ["A9031", "A9032", "A9033", "A9034", "A9035", "A9036", "A9030",
              "A3170", "A9037", "A9038", "A3160", "A3130", "A2040", "A2010",
              "A2013", "A3030", "A3010", "A9012", "A9011", "A9018", "2745",
              "A9039"]
emplois = ["A4012", "A3180", "A3170", "A3082", "A3084", "A3011", "A3012",
           "A3160", "A3161", "A3162", "A3131", "A3132", "A2041", "A2042",
           "A2011", "A2012", "A9010", "A9011", "A9012", "A9013", "A9014",
           "A9015", "A9016", "A9032", "A3030", "21810None", "21810"]



class transaction:
    def __init__(self, log, CEouBP, ligne=None, RouD=None, ws_banque=None):
        self.ligne = ligne
        self.CEouBP = CEouBP
        if (ligne, RouD, ws_banque) != (None, None, None):
            if RouD == "D":
                self.montant = ws_banque.cell(row=ligne, column=16).value
            elif RouD == "R":
                self.montant = ws_banque.cell(row=ligne, column=15).value
            else:
                print "ERREUR dans le constructeur de l'objet transaction"
                exit()
            self.RouE = ws_banque.cell(row=ligne, column=4).value
            self.code = str(ws_banque.cell(row=ligne, column=5).value) \
                + str(ws_banque.cell(row=ligne, column=6).value)
            self.antenne = ws_banque.cell(row=ligne, column=14).value
            self.piece = ws_banque.cell(row=ligne, column=7).value
            self.objet = ws_banque.cell(row=ligne, column=8).value
            self.libelle = ws_banque.cell(row=ligne, column=9).value
            self.nature = ws_banque.cell(row=ligne, column=10).value
            self.numero = ws_banque.cell(row=ligne, column=11).value
            self.financeur = ws_banque.cell(row=ligne, column=13).value
            self.date = ws_banque.cell(row=ligne, column=1).value
            if (RouD == "D" and self.RouE == "E") \
               or (RouD == "R" and self.RouE == "R"):
                self.regularisation = False
            else:
                self.regularisation = True
                log.debug("classe transaction: transaction.__init__ ")
                log.warning("--> regularisation en ligne : " + str(ligne) +
                            " du brouillard " + str(CEouBP))
                self.log_transaction_debug_light(log)
        else:
            self.ligne = 0
            self.RouD = "?"  # Recette ou Depense
            self.code = 0
            self.antenne = 0
            self.montant = 0

    def log_transaction_debug(self, log):
        log.debug("    self.ligne   = " + str(self.ligne))
        log.debug("    self.RouE    = " + str(self.RouE))
        log.debug("    self.code    = " + str(self.code))
        log.debug("    self.antenne = " + str(self.antenne))
        log.debug("    self.montant = " + str(self.montant))

    def log_transaction_debug_light(self, log):
        log.debug("    self.code = " + str(self.code))
        log.debug("    self.montant = " + str(self.montant))


class Balance:
    def __init__(self, file_name, log,
                 (liste_depense, liste_recette)=(None, None)):
        self.file_balance_output = \
            re.search('(\w+).xlsx', file_name).group(1) + "_new.xlsx"
        command = "cp " + file_name + " " + self.file_balance_output
        log.info(command)
        subprocess.call(command, shell=True)
        self.wb_out = openpyxl.load_workbook(self.file_balance_output,
                                             data_only=True)
        self.ws = self.wb_out.get_sheet_by_name('Balance')
        self.solde_initial = self.ws.cell(row=56, column=15).value
        self.solde_final = None
        self.resultat = None
        self.resultat_cumul = None
        self.date_initiale = self.ws.cell(row=1, column=18).value
        self.date_Format = self.ws.cell(row=1, column=18).number_format
        self.date_initiale_annee = int(re.search("^(20\d\d)-\d\d-.+",
                                       str(self.date_initiale)).group(1))
        self.date_initiale_mois = int(re.search("^20\d\d-([01]\d)-.+",
                                      str(self.date_initiale)).group(1))
        self.date_finale_mois = self.date_initiale_mois + 1
        _, date_dinale_LastDay = calendar.monthrange(self.date_initiale_annee,
                                                     self.date_finale_mois)
        self.date_finale = datetime.datetime(self.date_initiale_annee,
                                             self.date_finale_mois,
                                             date_dinale_LastDay)
        if True:
            ''' Depenses '''
            (self.A4012, self.A3180, self.A3170) = ([], [], [])
            (self.A3082, self.A3084, self.A3011) = ([], [], [])
            (self.A3012, self.A3160, self.A3161) = ([], [], [])
            (self.A3162, self.A3131, self.A3132) = ([], [], [])
            (self.A2041, self.A2042, self.A2042) = ([], [], [])
            (self.A2011, self.A2012, self.A9010) = ([], [], [])
            (self.A9011, self.A9012, self.A9013) = ([], [], [])
            (self.A9014, self.A9015, self.A9016) = ([], [], [])
            (self.A9032, self.A3030, self.A21810) = ([], [], [])

        if (liste_depense, liste_recette) != (None, None):
            self.ws.cell(row=1, column=18).value = self.date_finale
            self.ws.cell(row=1, column=18).number_format = self.date_Format
            self.ws.cell(row=55, column=13).value = self.date_finale
            self.ws.cell(row=55, column=13).number_format = self.date_Format
            self.ws.cell(row=56, column=13).value = self.date_finale
            self.ws.cell(row=56, column=13).number_format = self.date_Format
            log.info("NETTOYAGE DE LA BALANCE")
            self.nettoyage_balance()

            log.info("PEUPLEMENT DE LA BALANCE")
            log.debug("....DEPENSES")
            debug_antenne = None
            self.peuple_balance_depenses(liste_depense, log, debug_antenne)
            self.show_depenses_peuplement(debug_antenne)
            log.debug("....RECETTES")
            self.peuple_balance_recettes(liste_recette, log)
            log.debug("....SOUS TOTAUX DEPENSES")
            self.totaux_balance_depense()
            log.debug("....SOUS TOTAUX RECETTES")
            self.totaux_balance_recettes()
            log.debug("....TOTAUX")
            self.resultat = self.totaux_balance()

    def show_depenses_peuplement(self, debug_antenne):
        if not debug_antenne:
            return
        print "....debug antenne : " + str(debug_antenne)
        tmp = 0
        for elements in self.A3170:
            elements.imprime()
            tmp += elements.montant
        print "........total A3170 = " + str(tmp) + "\n"

        tmp = 0
        for elements in self.A3082:
            elements.imprime()
            tmp += elements.montant
        print "........total A3082 = " + str(tmp) + "\n"

        tmp = 0
        for elements in self.A3160:
            elements.imprime()
            tmp += elements.montant
        print "........total A3160 = " + str(tmp) + "\n"

        tmp = 0
        for elements in self.A3162:
            elements.imprime()
            tmp += elements.montant
        print "........total A3162 = " + str(tmp) + "\n"

        tmp = 0
        for elements in self.A2041:
            elements.imprime()
            tmp += elements.montant
        print "........total A2041 = " + str(tmp) + "\n"

        tmp = 0
        for elements in self.A2042:
            elements.imprime()
            tmp += elements.montant
        print "........total A2042 = " + str(tmp) + "\n"

        tmp = 0
        for elements in self.A2011:
            elements.imprime()
            tmp += elements.montant
        print "........total A2011 = " + str(tmp) + "\n"

        tmp = 0
        for elements in self.A9012:
            elements.imprime()
            tmp += elements.montant
        print "........total A9012 = " + str(tmp) + "\n"

    def nettoyage_balance(self):
        l_recettes = range(7, 14)+[16, 22, 24, 25, 26, 28, 31, 34, 36, 37, 38,
                                   43, 45, 47, 48, 49, 51, 52, 53]
        for ligne in l_recettes:
            for col in range(3, 11):
                self.ws.cell(row=ligne, column=col).value = 0
        l_depenses = [8, 9, 11, 12, 13, 14, 16, 17, 18, 20, 21, 22, 23, 25, 26,
                      27, 29, 30, 31, 33, 34, 35] + \
            range(37, 44)+[45]+range(47, 55)
        for ligne in l_depenses:
            for col in range(13, 21):
                self.ws.cell(row=ligne, column=col).value = 0

    def peuple_balance_depenses(self, liste, log, debug_antenne=None):
        '''
        - Peuple les depenses
        '''
        for i in range(0, len(liste)):
            try :
                column_antenne = int(antennes.index(liste[i].antenne)) + 13
            except:
                print "\n ERREUR : Antenne INCONNUE"
                log.critical("Antenne = " + str(liste[i].antenne))
                log.critical("brouillard = " + str(liste[i].CEouBP))
                log.critical("ligne du brouillard = " + str(liste[i].ligne))
                exit()
            self.peuple_balance_depenses_antenne(liste, i, column_antenne, log,
                                                 debug_antenne=debug_antenne)

    def peuple_balance_depenses_antenne(self, liste, i, antenne, log,
                                        debug_antenne=None):
        if liste[i].code == "A4012":
            if liste[i].antenne == debug_antenne:
                self.A4012.append(liste[i])
                pass
            self.ws.cell(row=8, column=antenne, value=liste[i].montant +
                         self.ws.cell(row=8, column=antenne).value)
        elif liste[i].code == "A3180":
            if liste[i].antenne == debug_antenne:
                self.A3180.append(liste[i])
                pass
            self.ws.cell(row=11, column=antenne, value=liste[i].montant +
                         self.ws.cell(row=11, column=antenne).value)
        elif liste[i].code == "A3170":
            if liste[i].antenne == debug_antenne:
                self.A3170.append(liste[i])
                pass
            self.ws.cell(row=12, column=antenne, value=liste[i].montant +
                         self.ws.cell(row=12, column=antenne).value)
        elif liste[i].code == "A3082":
            if liste[i].antenne == debug_antenne:
                self.A3082.append(liste[i])
                pass
            self.ws.cell(row=13, column=antenne, value=liste[i].montant +
                         self.ws.cell(row=13, column=antenne).value)
        elif liste[i].code == "A3084":
            if liste[i].antenne == debug_antenne:
                self.A3084.append(liste[i])
                pass
            self.ws.cell(row=14, column=antenne, value=liste[i].montant +
                         self.ws.cell(row=14, column=antenne).value)
        elif liste[i].code == "A3011":
            if liste[i].antenne == debug_antenne:
                self.A3011.append(liste[i])
                pass
            self.ws.cell(row=16, column=antenne, value=liste[i].montant +
                         self.ws.cell(row=16, column=antenne).value)
        elif liste[i].code == "A3012":
            if liste[i].antenne == debug_antenne:
                self.A3012.append(liste[i])
                pass
            self.ws.cell(row=17, column=antenne, value=liste[i].montant +
                         self.ws.cell(row=17, column=antenne).value)
        elif liste[i].code == "A3160":
            if liste[i].antenne == debug_antenne:
                self.A3160.append(liste[i])
                pass
            self.ws.cell(row=20, column=antenne, value=liste[i].montant +
                         self.ws.cell(row=20, column=antenne).value)
        elif liste[i].code == "A3161":
            if liste[i].antenne == debug_antenne:
                self.A3161.append(liste[i])
                pass
            self.ws.cell(row=21, column=antenne, value=liste[i].montant +
                         self.ws.cell(row=21, column=antenne).value)
        elif liste[i].code == "A3162":
            if liste[i].antenne == debug_antenne:
                self.A3162.append(liste[i])
                pass
            self.ws.cell(row=22, column=antenne, value=liste[i].montant +
                         self.ws.cell(row=22, column=antenne).value)
        elif liste[i].code == "A3131":
            if liste[i].antenne == debug_antenne:
                self.A3131.append(liste[i])
                pass
            self.ws.cell(row=25, column=antenne, value=liste[i].montant +
                         self.ws.cell(row=25, column=antenne).value)
        elif liste[i].code == "A3132":
            if liste[i].antenne == debug_antenne:
                self.A3132.append(liste[i])
                pass
            self.ws.cell(row=26, column=antenne, value=liste[i].montant +
                         self.ws.cell(row=26, column=antenne).value)
        elif liste[i].code == "A2041":
            if liste[i].antenne == debug_antenne:
                self.A2041.append(liste[i])
                pass
            self.ws.cell(row=29, column=antenne, value=liste[i].montant +
                         self.ws.cell(row=29, column=antenne).value)
        elif liste[i].code == "A2042":
            if liste[i].antenne == debug_antenne:
                self.A2042.append(liste[i])
                pass
            self.ws.cell(row=30, column=antenne, value=liste[i].montant +
                         self.ws.cell(row=30, column=antenne).value)
        elif liste[i].code == "A2011":
            if liste[i].antenne == debug_antenne:
                self.A2011.append(liste[i])
                pass
            self.ws.cell(row=33, column=antenne, value=liste[i].montant +
                         self.ws.cell(row=33, column=antenne).value)
        elif liste[i].code == "A2012":
            if liste[i].antenne == debug_antenne:
                self.A2012.append(liste[i])
                pass
            self.ws.cell(row=34, column=antenne, value=liste[i].montant +
                         self.ws.cell(row=34, column=antenne).value)
        elif liste[i].code == "A9010":
            if liste[i].antenne == debug_antenne:
                self.A9010.append(liste[i])
                pass
            self.ws.cell(row=37, column=antenne, value=liste[i].montant +
                         self.ws.cell(row=37, column=antenne).value)
        elif liste[i].code == "A9011":
            if liste[i].antenne == debug_antenne:
                self.A9011.append(liste[i])
                pass
            self.ws.cell(row=38, column=antenne, value=liste[i].montant +
                         self.ws.cell(row=38, column=antenne).value)
        elif liste[i].code == "A9012":
            if liste[i].antenne == debug_antenne:
                self.A9012.append(liste[i])
                pass
            self.ws.cell(row=39, column=antenne, value=liste[i].montant +
                         self.ws.cell(row=39, column=antenne).value)
        elif liste[i].code == "A9013":
            if liste[i].antenne == debug_antenne:
                self.A9013.append(liste[i])
                pass
            self.ws.cell(row=40, column=antenne, value=liste[i].montant +
                         self.ws.cell(row=40, column=antenne).value)
        elif liste[i].code == "A9014":
            if liste[i].antenne == debug_antenne:
                self.A9014.append(liste[i])
                pass
            self.ws.cell(row=41, column=antenne, value=liste[i].montant +
                         self.ws.cell(row=41, column=antenne).value)
        elif liste[i].code == "A9015":
            if liste[i].antenne == debug_antenne:
                self.A9015.append(liste[i])
                pass
            self.ws.cell(row=42, column=antenne, value=liste[i].montant +
                         self.ws.cell(row=42, column=antenne).value)
        elif liste[i].code == "A9016":
            if liste[i].antenne == debug_antenne:
                self.A9016.append(liste[i])
                pass
            self.ws.cell(row=43, column=antenne, value=liste[i].montant +
                         self.ws.cell(row=43, column=antenne).value)
        elif liste[i].code == "A9032":
            if liste[i].antenne == debug_antenne:
                self.A9032.append(liste[i])
                pass
            self.ws.cell(row=45, column=antenne, value=liste[i].montant +
                         self.ws.cell(row=45, column=antenne).value)
        elif liste[i].code == "A3030":
            if liste[i].antenne == debug_antenne:
                self.A3030.append(liste[i])
                pass
            self.ws.cell(row=47, column=antenne, value=liste[i].montant +
                         self.ws.cell(row=47, column=antenne).value)
        elif liste[i].code == "21810None" or liste[i].code == "21810":
            if liste[i].antenne == debug_antenne:
                self.A21810.append(liste[i])
                pass
            log.warning("   investissement en ligne : " + str(liste[i].ligne))
            self.ws.cell(row=52, column=antenne, value=liste[i].montant +
                         self.ws.cell(row=52, column=antenne).value)
        else:
            log.critical("\n ERREUR : Transaction depenses non traitee")
            log.critical("brouillard = " + str(liste[i].CEouBP))
            log.critical("code imputation = " + str(liste[i].code))
            log.critical("ligne = " + str(liste[i].ligne))
            exit()

    def peuple_balance_recettes(self, liste, log):
        '''
        - Peuple les depenses
        '''
        for i in range(0, len(liste)):
            try :
                column_antenne = int(antennes.index(liste[i].antenne)) + 3
            except:
                log.critical("\n ERREUR: Antenne INCONNUE dans \
                             methode peuple_balance_recettes")
                log.critical("Antenne = " + str(liste[i].antenne))
                log.critial("ligne du brouillard = " + str(liste[i].ligne))
                exit()
            self.peuple_balance_recettes_antenne(liste, i, column_antenne, log)

    def peuple_balance_recettes_antenne(self, liste, i, antenne, log):
        if liste[i].code == "A9031":
            val = liste[i].montant+self.ws.cell(row=7, column=antenne).value
            self.ws.cell(row=7, column=antenne, value=val)
        elif liste[i].code == "A9032":
            val = liste[i].montant+self.ws.cell(row=8, column=antenne).value
            self.ws.cell(row=8, column=antenne, value=val)
        elif liste[i].code == "A9033":
            val = liste[i].montant+self.ws.cell(row=9, column=antenne).value
            self.ws.cell(row=9, column=antenne, value=val)
        elif liste[i].code == "A9034":
            val = liste[i].montant+self.ws.cell(row=10, column=antenne).value
            self.ws.cell(row=10, column=antenne, value=val)
        elif liste[i].code == "A9035":
            val = liste[i].montant+self.ws.cell(row=11, column=antenne).value
            self.ws.cell(row=11, column=antenne, value=val)
        elif liste[i].code == "A9036":
            val = liste[i].montant+self.ws.cell(row=12, column=antenne).value
            self.ws.cell(row=12, column=antenne, value=val)
        elif liste[i].code == "A9030":
            val = liste[i].montant+self.ws.cell(row=13, column=antenne).value
            self.ws.cell(row=13, column=antenne, value=val)
        elif liste[i].code == "A3170":
            val = liste[i].montant+self.ws.cell(row=22, column=antenne).value
            self.ws.cell(row=22, column=antenne, value=val)
        elif liste[i].code == "A9037":
            val = liste[i].montant+self.ws.cell(row=24, column=antenne).value
            self.ws.cell(row=24, column=antenne, value=val)
        elif liste[i].code == "A9038":
            val = liste[i].montant+self.ws.cell(row=25, column=antenne).value
            self.ws.cell(row=25, column=antenne, value=val)
        elif liste[i].code == "A9039":
            val = liste[i].montant+self.ws.cell(row=26, column=antenne).value
            self.ws.cell(row=26, column=antenne, value=val)
        elif liste[i].code == "A3130":
            val = liste[i].montant+self.ws.cell(row=31, column=antenne).value
            self.ws.cell(row=31, column=antenne, value=val)
        elif liste[i].code == "A2040":
            val = liste[i].montant+self.ws.cell(row=34, column=antenne).value
            self.ws.cell(row=34, column=antenne, value=val)
        elif liste[i].code == "A2010":
            val = liste[i].montant+self.ws.cell(row=36, column=antenne).value
            self.ws.cell(row=36, column=antenne, value=val)
        elif liste[i].code == "A2013":
            val = liste[i].montant+self.ws.cell(row=37, column=antenne).value
            self.ws.cell(row=37, column=antenne, value=val)
        elif liste[i].code == "A3030":
            val = liste[i].montant+self.ws.cell(row=43, column=antenne).value
            self.ws.cell(row=43, column=antenne, value=val)
        elif liste[i].code == "A3010":
            val = liste[i].montant+self.ws.cell(row=45, column=antenne).value
            self.ws.cell(row=45, column=antenne, value=val)
        elif liste[i].code == "A9012":
            log.debug("méthode peuple_balance_recettes_antenne")
            log.warning("--> regularisation A9012 depense ligne " +
                        str(liste[i].ligne) + " brouillard " + liste[i].CEouBP)
            liste[i].log_transaction_debug_light(log)
            val = -liste[i].montant+self.ws.cell(row=39,
                                                 column=antenne+10).value
            self.ws.cell(row=39, column=antenne+10, value=val)
        elif liste[i].code == "A9014":
            log.debug("méthode peuple_balance_recettes_antenne")
            log.warning("--> regularisation A9014 depense ligne " +
                        str(liste[i].ligne) + " brouillard " + liste[i].CEouBP)
            liste[i].log_transaction_debug_light(log)
            val = -liste[i].montant+self.ws.cell(row=41,
                                                 column=antenne+10).value
            self.ws.cell(row=39, column=antenne+10, value=val)
        elif liste[i].code == "A9011" or liste[i].code == "A9018":
            log.warning("--> interets financiers en ligne : " +
                        str(liste[i].ligne))
            val = liste[i].montant + self.ws.cell(row=51, column=antenne).value
            self.ws.cell(row=51, column=antenne, value=val)
        elif str(liste[i].code) == "2745":
            log.warning("--> versement emprunt : " +
                        str(liste[i].ligne))
            val = liste[i].montant + self.ws.cell(row=51, column=antenne).value
            self.ws.cell(row=52, column=antenne, value=val)
        else:
            log.critical("\n ERREUR : Transaction recette non traitee dans methode" +\
                         " peuple_balance_recettes_antenne")
            log.critical("ligne du brouillard = " + str(liste[i].ligne))
            log.critical("brouillard = " + str(liste[i].CEouBP))
            log.critical("code imputation = " + str(liste[i].code))
            exit()

    def totaux_balance_depense(self):
        colonnes_antennes = [13, 14, 15, 16, 17, 18, 19]
        for antenne in colonnes_antennes:
            self.ws.cell(row=9, column=antenne,
                         value=self.ws.cell(row=8, column=antenne).value)
            self.ws.cell(row=18, column=antenne,
                         value=self.ws.cell(row=11, column=antenne).value +
                         self.ws.cell(row=12, column=antenne).value +
                         self.ws.cell(row=13, column=antenne).value +
                         self.ws.cell(row=14, column=antenne).value +
                         self.ws.cell(row=16, column=antenne).value +
                         self.ws.cell(row=17, column=antenne).value)
            self.ws.cell(row=23, column=antenne,
                         value=self.ws.cell(row=20, column=antenne).value +
                         self.ws.cell(row=21, column=antenne).value +
                         self.ws.cell(row=22, column=antenne).value)
            self.ws.cell(row=27, column=antenne,
                         value=self.ws.cell(row=25, column=antenne).value +
                         self.ws.cell(row=26, column=antenne).value)
            self.ws.cell(row=31, column=antenne,
                         value=self.ws.cell(row=29, column=antenne).value +
                         self.ws.cell(row=30, column=antenne).value)
            self.ws.cell(row=35, column=antenne,
                         value=self.ws.cell(row=33, column=antenne).value +
                         self.ws.cell(row=34, column=antenne).value)
            self.ws.cell(row=48, column=antenne,
                         value=self.ws.cell(row=37, column=antenne).value +
                         self.ws.cell(row=38, column=antenne).value +
                         self.ws.cell(row=39, column=antenne).value +
                         self.ws.cell(row=40, column=antenne).value +
                         self.ws.cell(row=41, column=antenne).value +
                         self.ws.cell(row=42, column=antenne).value +
                         self.ws.cell(row=43, column=antenne).value +
                         self.ws.cell(row=45, column=antenne).value +
                         self.ws.cell(row=47, column=antenne).value)
            self.ws.cell(row=49, column=antenne,
                         value=self.ws.cell(row=9, column=antenne).value +
                         self.ws.cell(row=18, column=antenne).value +
                         self.ws.cell(row=23, column=antenne).value +
                         self.ws.cell(row=27, column=antenne).value +
                         self.ws.cell(row=31, column=antenne).value +
                         self.ws.cell(row=35, column=antenne).value +
                         self.ws.cell(row=48, column=antenne).value)
            self.ws.cell(row=53, column=antenne,
                         value=self.ws.cell(row=49, column=antenne).value +
                         self.ws.cell(row=51, column=antenne).value +
                         self.ws.cell(row=52, column=antenne).value)

    def totaux_balance_recettes(self):
        column_antenne = [3, 4, 5, 6, 7, 8, 9]
        for antenne in column_antenne:
            self.ws.cell(row=16, column=antenne,
                         value=self.ws.cell(row=7, column=antenne).value +
                         self.ws.cell(row=8, column=antenne).value +
                         self.ws.cell(row=9, column=antenne).value +
                         self.ws.cell(row=10, column=antenne).value +
                         self.ws.cell(row=11, column=antenne).value +
                         self.ws.cell(row=12, column=antenne).value +
                         self.ws.cell(row=13, column=antenne).value)
            self.ws.cell(row=28, column=antenne,
                         value=self.ws.cell(row=22, column=antenne).value +
                         self.ws.cell(row=24, column=antenne).value +
                         self.ws.cell(row=25, column=antenne).value +
                         self.ws.cell(row=26, column=antenne).value)
            self.ws.cell(row=38, column=antenne,
                         value=self.ws.cell(row=36, column=antenne).value +
                         self.ws.cell(row=37, column=antenne).value)
            self.ws.cell(row=47, column=antenne,
                         value=self.ws.cell(row=43, column=antenne).value +
                         self.ws.cell(row=45, column=antenne).value)
            self.ws.cell(row=49, column=antenne,
                         value=self.ws.cell(row=16, column=antenne).value +
                         self.ws.cell(row=28, column=antenne).value +
                         self.ws.cell(row=31, column=antenne).value +
                         self.ws.cell(row=34, column=antenne).value +
                         self.ws.cell(row=38, column=antenne).value +
                         self.ws.cell(row=47, column=antenne).value)
            self.ws.cell(row=53, column=antenne,
                         value=self.ws.cell(row=49, column=antenne).value +
                         self.ws.cell(row=51, column=antenne).value +
                         self.ws.cell(row=52, column=antenne).value)

    def totaux_balance(self):
        l_recettes = range(7, 13)+[16, 22, 24, 25, 26, 28, 31, 34, 36, 37, 38,
                                   43, 45, 47, 48, 49, 51, 52, 53]
        for ligne in l_recettes:
            self.ws.cell(row=ligne, column=10,
                         value=self.ws.cell(row=ligne, column=3).value +
                         self.ws.cell(row=ligne, column=4).value +
                         self.ws.cell(row=ligne, column=5).value +
                         self.ws.cell(row=ligne, column=6).value +
                         self.ws.cell(row=ligne, column=7).value +
                         self.ws.cell(row=ligne, column=8).value +
                         self.ws.cell(row=ligne, column=9).value)
        c_resultat = range(3, 10)
        for colonne in c_resultat:
            self.ws.cell(row=50, column=colonne+10,
                         value=self.ws.cell(row=49, column=colonne).value -
                         self.ws.cell(row=49, column=colonne+10).value)
            self.ws.cell(row=54, column=colonne+10,
                         value=self.ws.cell(row=53, column=colonne).value -
                         self.ws.cell(row=53, column=colonne+10).value)
        l_depenses = [8, 9]+range(11, 15)+[16, 17, 18, 21, 22, 23, 25, 26, 27,
                                           29, 30, 31, 33, 34, 35] + \
            range(37, 44)+[45]+range(47, 55)
        for ligne in l_depenses:
            self.ws.cell(row=ligne, column=20,
                         value=self.ws.cell(row=ligne, column=13).value +
                         self.ws.cell(row=ligne, column=14).value +
                         self.ws.cell(row=ligne, column=15).value +
                         self.ws.cell(row=ligne, column=16).value +
                         self.ws.cell(row=ligne, column=17).value +
                         self.ws.cell(row=ligne, column=18).value +
                         self.ws.cell(row=ligne, column=19).value)
        self.resultat = self.ws.cell(row=54, column=20).value
        return self.resultat

    def cumul(self, log, (solde_banque_CE, solde_banque_BP)=(None, None)):
        print "CUMUL DEPUIS JANVIER"
        l_recettes = range(7, 13)+[16, 22, 24, 25, 26, 28, 31, 34, 36, 37, 38,
                                   43, 45, 47, 48, 49, 51, 52, 53]
        for ligne in l_recettes:
            for col in range(26, 34):
                self.ws.cell(row=ligne, column=col).value = \
                    self.ws.cell(row=ligne, column=col).value + \
                    self.ws.cell(row=ligne, column=col-23).value
        l_depenses = [8, 9]+range(11, 15)+range(16, 19)+range(21, 24) + \
            [25, 26, 27, 29, 30, 31, 33, 34, 35] + \
            range(37, 44)+[45]+range(47, 55)
        for ligne in l_depenses:
            for col in range(36, 44):
                self.ws.cell(row=ligne, column=col).value = \
                    self.ws.cell(row=ligne, column=col).value + \
                    self.ws.cell(row=ligne, column=col-23).value

        self.resultat_cumul = self.ws.cell(row=54, column=43).value

        self.ws.cell(row=1, column=41).value = self.date_finale
        self.ws.cell(row=1, column=41).number_format = self.date_Format
        for ligne in [55, 56]:
            self.ws.cell(row=ligne, column=36).value = self.date_finale
            self.ws.cell(row=ligne, column=36).number_format = self.date_Format
        if (solde_banque_CE, solde_banque_BP) is not (None, None):
            self.solde_final = self.solde_initial + self.resultat
            self.ws.cell(row=56, column=15, value=self.solde_final)
            self.ws.cell(row=56, column=38, value=self.solde_final)
            if solde_banque_CE is not None and solde_banque_BP is not None:
                self.ws.cell(row=55, column=15,
                             value=solde_banque_CE + solde_banque_BP)
                self.ws.cell(row=55, column=38,
                             value=solde_banque_CE + solde_banque_BP)
            elif solde_banque_CE is not None:
                self.ws.cell(row=55, column=15, value=solde_banque_CE)
                self.ws.cell(row=55, column=38, value=solde_banque_CE)
            elif solde_banque_BP is not None:
                self.ws.cell(row=55, column=15, value=solde_banque_BP)
                self.ws.cell(row=55, column=38, value=solde_banque_BP)
            else:
                print "\nERREUR dans balance.cumul() fonction"
                print "Quitting"
                exit()


class brouillard:
    def __init__(self, filename, CEouBP):
        ''' "CE" ou "BP" '''
        self.CEouBP = CEouBP
        self.wb = openpyxl.load_workbook(filename, data_only=True)
        self.ws = self.wb.get_sheet_by_name('Brouillard')

    def decipher(self, log, NDI_exist=True):
        '''
        - Get all transactions from brouillard CE
        '''
        log.info("DECHIFFREMENT DU BROUILLARD DE LA BANQUE " + self.CEouBP)
        liste_depenses = []
        liste_recettes = []
        liste_NDI_depenses = []
        liste_NDI_recettes = []
        i = 1
        # DEPENSES et RECETTES
        if self.CEouBP == "CE":  # le brouillard BP n'a que des NDI
            while self.ws.cell(row=i, column=1).value != "DEPENSES":
                i += 1
            i_premiere_depense = i + 1
            log.debug("....Les depenses commencent en ligne " +
                      str(i_premiere_depense))
            i = i_premiere_depense
            while self.ws.cell(row=i, column=1).value != "RECETTES":
                T = transaction(log, self.CEouBP, i, "D", self.ws)
                liste_depenses.append(T)
                i += 1
            i_derniere_depense = i - 1
            log.debug("....Les depenses terminent en ligne " +
                      str(i_derniere_depense))
            i_premiere_recette = i + 1
            log.debug("....Les recettes commencent en ligne " +
                      str(i_premiere_recette))
            i = i_premiere_recette
            while re.match("(^20\d\d-\d\d-\d\d ).+",
                           str(self.ws.cell(row=i, column=1).value)):
                T = transaction(log, self.CEouBP, i, "R", self.ws)
                liste_recettes.append(T)
                i += 1
            i_derniere_recettes = i - 1
            log.debug("....Les recettes terminent en ligne " +
                      str(i_derniere_recettes))
        elif self.CEouBP == "BP":
            while not re.match("(NDI D).*",
                               str(self.ws.cell(row=i, column=1).value)):
                i += 1
        else:
            log.critical("Probleme d'identification du brouillard")
            exit()
        if not NDI_exist:
            self.depenses = liste_depenses
            self.recettes = liste_recettes
            self.NDI_depenses = liste_NDI_depenses
            self.NDI_recettes = liste_NDI_recettes
            return
        # NDI
        i_premiere_NDI_R = i + 1
        log.debug("....Les NDI depenses " + self.CEouBP +
                  " commencent en ligne " + str(i_premiere_NDI_R))
        i = i_premiere_NDI_R
        while not re.match("(NDI R).*",
                           str(self.ws.cell(row=i, column=1).value)):
            T = transaction(log, self.CEouBP, i, "D", self.ws)
            liste_NDI_depenses.append(T)
            i += 1
        i_derniere_NDI_D = i - 1
        log.debug("....Les NDI depenses " + self.CEouBP +
                  " terminent en ligne " + str(i_derniere_NDI_D))
        i_premiere_NDI_R = i + 1
        log.debug("....Les NDI recettes " + self.CEouBP +
                  " commencent en ligne " + str(i_premiere_NDI_R))
        i = i_premiere_NDI_R
        while re.match("^\d{4}\D\d{2}\D\d{2}\s.*",
                       str(self.ws.cell(row=i, column=1).value)):
            T = transaction(log, self.CEouBP, i, "R", self.ws)
            liste_NDI_recettes.append(T)
            i += 1
        i_derniere_NDI_R = i - 1
        log.debug("....Les NDI recettes " + self.CEouBP +
                  " terminent en ligne " + str(i_derniere_NDI_R))
        log.debug("....Le brouillard de la banque " + self.CEouBP +
                  " a ete dechifre sans trouver d'erreurs")
        self.depenses = liste_depenses
        self.recettes = liste_recettes
        self.NDI_depenses = liste_NDI_depenses
        self.NDI_recettes = liste_NDI_recettes

    def resultat(self, liste_depenses, liste_recettes,
                 antenne=None, Code=None):
        resultat = 0
        if liste_recettes:
            for i in range(0, len(liste_recettes)):
                if liste_recettes[i].antenne == antenne or not antenne:
                    if liste_recettes[i].code == Code or not Code:
                        resultat += liste_recettes[i].montant
        if liste_depenses:
            for i in range(0, len(liste_depenses)):
                if liste_depenses[i].antenne == antenne or not antenne:
                    if liste_depenses[i].code == Code or not Code:
                        resultat -= liste_depenses[i].montant
        return resultat


def debug_brou2bal((brouillard_CE_resulat, brouillard_BP_resulat),
                   brouillard, balance):
    print "\n==================  DEBUG  ======================="
    print "==================   " + brouillard.CEouBP + \
        "    ======================="
    if brouillard.CEouBP == "CE":
        print "Resultat inscrit sur le brouillard   = " + \
            str(brouillard_CE_resulat)
    elif brouillard.CEouBP == "BP":
        print "Resultat inscrit sur le brouillard   = " + \
            str(brouillard_BP_resulat)
    else:
        print "ERREUR dans la fonction debug"
        exit()
    print "Resultat balance                     = "+str(balance.resultat)
    print "----      Dechiffrage du brouillard         ----"
    print "depenses     = " + str(brouillard.resultat(
        brouillard.depenses, None))
    print "recettes     = " + str(brouillard.resultat(
        None, brouillard.recettes))
    print "NDI_depenses = " + str(brouillard.resultat(
        brouillard.NDI_depenses, None))
    print "NDI_recettes = " + str(brouillard.resultat(
        None, brouillard.NDI_recettes))
    print "TOTAL RECETTES = " + str(brouillard.resultat(
        None, brouillard.recettes + brouillard.NDI_recettes))
    print "TOTAL DEPENSES = " + str(brouillard.resultat(
        brouillard.depenses + brouillard.NDI_depenses, None))
    print "Resultat dechiffrage brouillard       = " + \
        str(brouillard.resultat(brouillard.depenses + brouillard.NDI_depenses,
                                brouillard.recettes + brouillard.NDI_recettes))
    print "---- Dechiffrage du brouillard CE par antenne ----"
    print "Sous total Recettes CE : 3969 = " + \
        str(brouillard.resultat(None, brouillard.recettes +
            brouillard.NDI_recettes, antenne=3969))
    print "Sous total Recettes CE : 4010 = " + \
        str(brouillard.resultat(None, brouillard.recettes +
            brouillard.NDI_recettes, antenne=4010))
    print "Sous total Recettes CE : 4011 = " + \
        str(brouillard.resultat(None, brouillard.recettes +
            brouillard.NDI_recettes, antenne=4011))
    print "Sous total Recettes CE : 4012 = " + \
        str(brouillard.resultat(None, brouillard.recettes +
            brouillard.NDI_recettes, antenne=4012))
    print "Sous total Recettes CE : 4013 = " + \
        str(brouillard.resultat(None, brouillard.recettes +
            brouillard.NDI_recettes, antenne=4013))
    print "Sous total Recettes CE : 4015 = " + \
        str(brouillard.resultat(None, brouillard.recettes +
            brouillard.NDI_recettes, antenne=4015))
    print "Sous total Recettes CE : 4016 = " + \
        str(brouillard.resultat(None, brouillard.recettes +
            brouillard.NDI_recettes, antenne=4016))
    print "Sous total Depenses CE : 3969 = " + \
        str(brouillard.resultat(brouillard.depenses +
            brouillard.NDI_depenses, None, antenne=3969))
    print "Sous total Depenses CE : 4010 = " + \
        str(brouillard.resultat(brouillard.depenses +
            brouillard.NDI_depenses, None, antenne=4010))
    print "Sous total Depenses CE : 4011 = " + \
        str(brouillard.resultat(brouillard.depenses +
            brouillard.NDI_depenses, None, antenne=4011))
    print "Sous total Depenses CE : 4012 = " + \
        str(brouillard.resultat(brouillard.depenses +
            brouillard.NDI_depenses, None, antenne=4012))
    print "Sous total Depenses CE : 4013 = " + \
        str(brouillard.resultat(brouillard.depenses +
            brouillard.NDI_depenses, None, antenne=4013))
    print "Sous total Depenses CE : 4015 = " + \
        str(brouillard.resultat(brouillard.depenses +
            brouillard.NDI_depenses, None, antenne=4015))
    print "Sous total Depenses CE : 4016 = " + \
        str(brouillard.resultat(brouillard.depenses +
            brouillard.NDI_depenses, None, antenne=4016))
    print "Sous total Resultat CE : 3969 = " + \
        str(brouillard.resultat(brouillard.depenses + brouillard.NDI_depenses,
            brouillard.recettes + brouillard.NDI_recettes, antenne=3969))
    print "Sous total Resultat CE : 4010 = " + \
        str(brouillard.resultat(brouillard.depenses + brouillard.NDI_depenses,
            brouillard.recettes + brouillard.NDI_recettes, antenne=4010))
    print "Sous total Resultat CE : 4011 = " + \
        str(brouillard.resultat(brouillard.depenses + brouillard.NDI_depenses,
            brouillard.recettes + brouillard.NDI_recettes, antenne=4011))
    print "Sous total Resultat CE : 4012 = " + \
        str(brouillard.resultat(brouillard.depenses + brouillard.NDI_depenses,
            brouillard.recettes + brouillard.NDI_recettes, antenne=4012))
    print "Sous total Resultat CE : 4013 = " + \
        str(brouillard.resultat(brouillard.depenses + brouillard.NDI_depenses,
            brouillard.recettes + brouillard.NDI_recettes, antenne=4013))
    print "Sous total Resultat CE : 4015 = " + \
        str(brouillard.resultat(brouillard.depenses + brouillard.NDI_depenses,
            brouillard.recettes + brouillard.NDI_recettes, antenne=4015))
    print "Sous total Resultat CE : 4016 = " + \
        str(brouillard.resultat(brouillard.depenses + brouillard.NDI_depenses,
            brouillard.recettes + brouillard.NDI_recettes, antenne=4016))
    print "===============   END DEBUG  ====================="


def ajout_logger():
    """ Ajoute un logger
    Un formatteur ecrit sur la console
    Un formatteur ecrit dans activity.log
    Un formatteur pourrait envoyer des emails
    http://sametmax.com/ecrire-des-logs-en-python/
    """
    _logger = logging.getLogger()
    _logger.setLevel(logging.DEBUG)

    formatter = \
        logging.Formatter('%(asctime)s :: %(levelname)s :: %(message)s')
    file_handler = RotatingFileHandler('activity.log', 'a', 1000000, 1)
    file_handler.setLevel(logging.DEBUG)
    file_handler.setFormatter(formatter)
    _logger.addHandler(file_handler)

    steam_handler = logging.StreamHandler()
    steam_handler.setLevel(logging.INFO)
    _logger.addHandler(steam_handler)
    _logger.info("\nSTART")
    return _logger


def create_balance_DEBUG(config, brouillard, logger, CEouBP="CE"):
    if CEouBP == "CE":
        balance_CE = \
            Balance(config.get('Balance', 'file_in'), logger,
                    (brouillard.depenses + brouillard.NDI_depenses,
                     brouillard.recettes + brouillard.NDI_recettes))
        balance_CE.cumul(logger, (config.getfloat('Brouillard_CE',
                                                  'solde_bancaire'), None))
        print "enregistrement de la balance : Balance_CE_DEBUG.xlsx"
        balance_CE.wb_out.save("Balance_CE_DEBUG.xlsx")
        debug_brou2bal((config.getfloat('Brouillard_CE', 'resultat'),
                        None),
                       brouillard, balance_CE)
        print "Resultat balance CE = " + str(balance_CE.resultat)
        print "Resultat brouillards CE = " + \
            str(config.getfloat('Brouillard_CE', 'resultat')) + "\n\n"
    if CEouBP == "BP":
        brouillard_BP = brouillard(config.get('Brouillard_BP', 'file'), "BP")
        brouillard_BP.decipher(logger)
        if config.getboolean('Brouillard_BP', 'debug'):
            balance_BP = \
                Balance(config.get('Balance', 'file_in'), logger,
                        (brouillard_BP.depenses + brouillard_BP.NDI_depenses,
                         brouillard_BP.recettes + brouillard_BP.NDI_recettes))
            balance_CE.cumul(logger, (config.getfloat('Brouillard_BP',
                                                      'solde_bancaire'), None))
            print "enregistrement de la balance : Balance_BP_DEBUG.xlsx"
            balance_BP.wb_out.save("Balance_BP_DEBUG.xlsx")
            debug_brou2bal((None, config.getfloat('Brouillard_BP',
                                                  'resultat')),
                           brouillard_BP, balance_BP)
            print "Resultat balance BP = " + str(balance_BP.resultat)
            print "Resultat brouillards BP = " + \
                str(config.getfloat('Brouillard_BP', 'resultat')) + "\n\n"


def create_balance(config, brouillard_CE, brouillard_BP, logger):
        balance = \
            Balance(config.get('Balance', 'file_in'), logger,
                    (brouillard_CE.depenses + brouillard_CE.NDI_depenses +
                     brouillard_BP.depenses + brouillard_BP.NDI_depenses,
                     brouillard_CE.recettes + brouillard_CE.NDI_recettes +
                     brouillard_BP.recettes + brouillard_BP.NDI_recettes))
        balance.cumul(logger, (config.getfloat('Brouillard_CE',
                                               'solde_bancaire'),
                               config.getfloat('Brouillard_BP',
                                               'solde_bancaire')))
        balance.wb_out.save(balance.file_balance_output)

        logger.warning("=============== VERIFICATION =====================")

        logger.warning("Resultat balance CE+BP     = " + str(balance.resultat))
        logger.warning("Resultat brouillards CE+BP = " +
                       str(config.getfloat('Brouillard_CE', 'resultat') +
                           config.getfloat('Brouillard_BP', 'resultat')))
        logger.warning("Resultats cumules balance CE+BP     = " +
                       str(balance.resultat_cumul))
        logger.warning("Resultats cumules brouillards CE+BP = " +
                       str(config.getfloat('Brouillard_CE',
                                           'resultat_cumul') +
                           config.getfloat('Brouillard_BP',
                                           'resultat_cumul')))
        logger.warning("=============== INFORMATIONS =====================")
        logger.warning("Solde comptable CE+BP = " +
                       str(config.getfloat('Brouillard_CE',
                                           'solde_comptable') +
                           config.getfloat('Brouillard_BP',
                                           'solde_comptable')))
        logger.warning("Solde bancaire CE+BP = " +
                       str(config.getfloat('Brouillard_CE',
                                           'solde_bancaire') +
                           config.getfloat('Brouillard_BP',
                                           'solde_bancaire')))
        logger.warning("==================================================")


def create_fichier_detailoperation(log, config, depenses, recettes):
    log.info('CREATION DU FICHIER DE DEPENSES ET RECETTES')
    # nettoyer excel
    filename = config.get('DepensesRecettes', 'file_in')
    wb = openpyxl.load_workbook(filename, data_only=True)
    ws = wb.get_sheet_by_name('Feuil1')
    for l in range(9, 1001):
        for c in range(1, 19):
            ws.cell(row=l, column=c).value = None
    '''
    Pour chaque antenne
        pour chaque emplois
            pour toutes les recettes
                ecrire la recette si egale a l'emloi et l'antenne
                additionner la recette
            sauter une ligne
        sauter une ligne
    Pour chaque antenne
        pour chaque emplois
            pour toutes les depenses
                ecrire la depense si egale a la ressource et l'antenne
                additionner la depense
            sauter une ligne
        sauter une ligne
    Comparer depenses et recettes additionnées avec celles du brouillard
    '''
    ligne = 9
    wrote = False
    total_code = 0
    total_antenne = 0
    total_depenses = 0
    for antenne in antennes:
        for emploi in emplois:
            for depense in depenses:
                if depense.antenne == antenne and depense.code == emploi:
                    _tmp = ws.cell(row=ligne, column=1).number_format
                    ws.cell(row=ligne, column=1).value = depense.date
                    ws.cell(row=ligne, column=1).number_format = _tmp
                    ws.cell(row=ligne, column=5).value = depense.code
                    ws.cell(row=ligne, column=7).value = depense.piece
                    ws.cell(row=ligne, column=8).value = depense.objet
                    ws.cell(row=ligne, column=9).value = depense.libelle
                    ws.cell(row=ligne, column=10).value = depense.nature
                    ws.cell(row=ligne, column=11).value = depense.numero
                    ws.cell(row=ligne, column=12).value = depense.code
                    ws.cell(row=ligne, column=13).value = depense.financeur
                    ws.cell(row=ligne, column=14).value = depense.antenne
                    ws.cell(row=ligne, column=16).value = depense.montant
                    total_code += depense.montant
                    total_antenne += depense.montant
                    if depense.regularisation:
                        log.warning("--> regularisation ligne " +
                                    str(depense.ligne) + " brouillard " +
                                    str(depense.CEouBP))
                        ws.cell(row=ligne, column=19).value = "regularisation"
                    wrote = True
                    ligne += 1
            if wrote:
                wrote = False
                ws.cell(row=ligne-1, column=17).value = total_code
                total_code = 0
                ligne += 1
        ws.cell(row=ligne-2, column=18).value = total_antenne
        total_depenses += total_antenne
        total_antenne = 0
        ligne += 1

    ws.cell(row=ligne, column=1).value = "RECETTES"
    ws.cell(row=ligne, column=1).font  = Font(bold=True, name='Arial', size=10)
    ligne += 1
    wrote = False
    total_code = 0
    total_antenne = 0
    total_recettes = 0
    for antenne in antennes:
        for ressource in ressources:
            for recette in recettes:
                if recette.antenne == antenne and recette.code == ressource:
                    _tmp = ws.cell(row=ligne, column=1).number_format
                    ws.cell(row=ligne, column=1).value = recette.date
                    ws.cell(row=ligne, column=1).number_format = _tmp
                    ws.cell(row=ligne, column=5).value = recette.code
                    ws.cell(row=ligne, column=7).value = recette.piece
                    ws.cell(row=ligne, column=8).value = recette.objet
                    ws.cell(row=ligne, column=9).value = recette.libelle
                    ws.cell(row=ligne, column=10).value = recette.nature
                    ws.cell(row=ligne, column=11).value = recette.numero
                    ws.cell(row=ligne, column=13).value = recette.financeur
                    ws.cell(row=ligne, column=14).value = recette.antenne
                    ws.cell(row=ligne, column=15).value = recette.montant
                    total_code += recette.montant
                    total_antenne += recette.montant
                    if recette.regularisation:
                        log.warning("--> regularisation ligne " +
                                    str(recette.ligne) + " brouillard " +
                                    str(recette.CEouBP))
                        ws.cell(row=ligne, column=19).value = "regularisation"
                    wrote = True
                    ligne += 1
            if wrote:
                wrote = False
                ws.cell(row=ligne-1, column=17).value = total_code
                total_code = 0
                ligne += 1
        ws.cell(row=ligne-2, column=18).value = total_antenne
        total_recettes += total_antenne
        total_antenne = 0
        ligne += 1
    wb.save("DetailsDepensesRecettes_DEBUG.xlsx")
    log.info("Total depenses = " + str(total_depenses))
    log.info("Total recettes = " + str(total_recettes))
    log.info("Total resultat = " + str(total_recettes - total_depenses))



if __name__ == '__main__':
    _logger = ajout_logger()
    _config = ConfigParser.RawConfigParser()
    _config.read('brouillards2balance.cfg')

    if _config.getboolean('Balance', 'include_brouillard_CE'):
        _brouillard_CE = brouillard(_config.get('Brouillard_CE', 'file'), "CE")
        _brouillard_CE.decipher(_logger, _config.getboolean('Brouillard_CE',
                                                            'NDI_exist'))
        if _config.getboolean('Brouillard_CE', 'debug'):
            create_balance_DEBUG(_config, _brouillard_CE, _logger, "CE")

    if _config.getboolean('Balance', 'include_brouillard_BP'):
        _brouillard_BP = brouillard(_config.get('Brouillard_BP', 'file'), "BP")
        _brouillard_BP.decipher(_logger)
        if _config.getboolean('Brouillard_BP', 'debug'):
            create_balance_DEBUG(_config, _brouillard_BP, _logger, "BP")

    if _config.getboolean('Balance', 'balance_global_create'):
        create_balance(_config, _brouillard_CE, _brouillard_BP, _logger)

    if _config.getboolean('DepensesRecettes', 'faire'):
        if _config.getboolean('DepensesRecettes', 'detail_brouillard_CE'):
            create_fichier_detailoperation(_logger, _config, \
                    _brouillard_CE.depenses + _brouillard_CE.NDI_depenses, \
                    _brouillard_CE.recettes + _brouillard_CE.NDI_recettes)
        if _config.getboolean('DepensesRecettes', 'detail_brouillard_BP'):
            create_fichier_detailoperation(_logger, _config, \
                    _brouillard_CE.depenses + _brouillard_CE.NDI_depenses +\
                    _brouillard_BP.depenses + _brouillard_BP.NDI_depenses, \
                    _brouillard_CE.recettes + _brouillard_CE.NDI_recettes +\
                    _brouillard_CE.recettes + _brouillard_CE.NDI_recettes)
