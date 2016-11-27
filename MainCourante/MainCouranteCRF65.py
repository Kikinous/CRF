#!/usr/bin/env python
# -*- coding: UTF-8 -*-
#
# Main Courante
# Croix rouge Francaise 65
# 25 Nov 2016
# Julien Borghetti

import time
import logging
from logging.handlers import RotatingFileHandler
import openpyxl
from gui_victime import *
from gui_MainCourante import *
#import pdb ; pdb.set_trace()


class Victime:
    def __init__(self):
        self.Dossier_num     = None
        self.Arrivee_HH      = None
        self.Arrivee_MM      = None
        self.Origine         = None
        self.Destination     = None
        self.Nom             = None
        self.Prenom          = None
        self.Circonstances   = None
        self.Traitement      = None
        self.Numero_Bilan    = None
        self.Numero_Decharge = None
        self.Depart_HH       = None
        self.Depart_MM       = None
    def Set_defaut(self, numero):
        self.Dossier_num     = numero
        self.Arrivee_HH      = str(time.localtime()[3])
        self.Arrivee_MM      = str(time.localtime()[4])
        self.Origine         = ""
        self.Destination     = ""
        self.Nom             = ""
        self.Prenom          = ""
        self.Circonstances   = ""
        self.Traitement      = ""
        self.Depart_HH       = ""
        self.Depart_MM       = ""
    def Set_paul_durant(self, numero):
        self.Dossier_num     = numero
        self.Arrivee_HH      = "17"
        self.Arrivee_MM      = "03"
        self.Origine         = "signaleur"
        self.Destination     = "L.S.P."
        self.Nom             = "Durant"
        self.Prenom          = "Paul"
        self.Circonstances   = "Tombe a velo"
        self.Traitement      = "Desinfection et pansement"
        self.Numero_Bilan    = ""
        self.Numero_Decharge = ""
        self.Depart_HH       = "17"
        self.Depart_MM       = "10"
    def afficher(self):
        logger.debug(self.Arrivee_MM)
        logger.debug(self.Origine)
        logger.debug(self.Destination)
        logger.debug(self.Nom)
        logger.debug(self.Prenom)
        logger.debug(self.Circonstances)
        logger.debug(self.Traitement)
        logger.debug(self.Numero_Bilan)
        logger.debug(self.Numero_Decharge)
        logger.debug(self.Depart_HH)
        logger.debug(self.Depart_MM)


class GUI_Victime_Edition(GUI_Victime):
    def __init__(self, _victime_gui):
        logger.debug("Creation d'une nouvelle fiche pour : " + str(_victime_gui.Nom))
        GUI_Victime.__init__(self, None, wx.ID_ANY, "")
        self.victime_gui = _victime_gui
        self.montrer_victime()
    def enregistrer_fermer_gui_victime(self, event):  # wxGlade: GUI_Victime.<event_handler>
        self.enregistrer_victime()
        self.Close()
        event.Skip()
    def enregistrer_victime(self):
        logger.debug("enregistrer_victime()")
        self.victime_gui.Arrivee_HH      = self.text_ctrl_Arrivee_HH.GetValue()
        self.victime_gui.Arrivee_MM      = self.text_ctrl_Arrivee_MM.GetValue()
        self.victime_gui.Origine         = self.text_ctrl_origine.GetValue()
        self.victime_gui.Destination     = self.text_ctrl_destination.GetValue()
        self.victime_gui.Nom             = self.text_ctrl_Nom.GetValue()
        self.victime_gui.Prenom          = self.text_ctrl_Prenom.GetValue()
        self.victime_gui.Circonstances   = self.text_ctrl_Circonstances.GetValue()
        self.victime_gui.Traitement      = self.text_ctrl_Traitement.GetValue()
        self.victime_gui.Numero_Bilan    = self.text_ctrl_Numero_Bilan.GetValue()
        self.victime_gui.Numero_Decharge = self.text_ctrl_Numero_Decharge.GetValue()
        self.victime_gui.Depart_HH       = self.text_ctrl_Depart_HH.GetValue()
        self.victime_gui.Depart_MM       = self.text_ctrl_Depart_MM.GetValue()
        self.victime_gui.afficher()
    def montrer_victime(self):
        logger.debug("montrer_victime()")
        self.text_ctrl_Arrivee_HH.SetValue(str(self.victime_gui.Arrivee_HH))
        self.text_ctrl_Arrivee_MM.SetValue(str(self.victime_gui.Arrivee_MM))
        self.text_ctrl_origine.SetValue(self.victime_gui.Origine)
        self.text_ctrl_destination.SetValue(self.victime_gui.Destination)
        self.text_ctrl_Nom.SetValue(self.victime_gui.Nom)
        self.text_ctrl_Prenom.SetValue(self.victime_gui.Prenom)
        self.text_ctrl_Circonstances.SetValue(self.victime_gui.Circonstances)
        self.text_ctrl_Traitement.SetValue(self.victime_gui.Traitement)
        self.text_ctrl_Numero_Bilan.SetValue(str(self.victime_gui.Numero_Bilan))
        self.text_ctrl_Numero_Decharge.SetValue(str(self.victime_gui.Numero_Decharge))
        self.text_ctrl_Depart_HH.SetValue(str(self.victime_gui.Depart_HH))
        self.text_ctrl_Depart_MM.SetValue(str(self.victime_gui.Depart_MM))


class GUI_MainCourante_Fenetre(MainCourante_Fenetre):
    def __init__(self):
        MainCourante_Fenetre.__init__(self, None, wx.ID_ANY, "")
    def MetAJour_Victime(self, _victime, ligne):
        self.list_ctrl_ListeVictime.DeleteItem(ligne)
        logger.debug(_victime.Arrivee_HH)
        self.Insere_Victime(_victime,ligne)
    def Insere_Victime(self, _victime,ligne):
        _index = self.list_ctrl_ListeVictime.InsertStringItem(ligne, _victime.Arrivee_HH)
        logger.debug(_index)
        self.list_ctrl_ListeVictime.SetStringItem(_index, 1, _victime.Arrivee_MM )
        self.list_ctrl_ListeVictime.SetStringItem(_index, 2, _victime.Nom )
        self.list_ctrl_ListeVictime.SetStringItem(_index, 3, _victime.Prenom )
        self.list_ctrl_ListeVictime.SetStringItem(_index, 4, _victime.Circonstances )
        self.list_ctrl_ListeVictime.SetStringItem(_index, 5, _victime.Depart_HH )
        self.list_ctrl_ListeVictime.SetStringItem(_index, 6, _victime.Depart_MM )
    def Nouvelle(self, event):  # wxGlade: MainCourante_Fenetre.<event_handler>
        logger.debug("Event handler 'Nouvelle'")
        gettext.install("DialogVictime")
        DialogVictime = wx.App()
        Liste_Victimes.append(Victime())
        Liste_Victimes[len(Liste_Victimes)-1].Set_defaut(str(len(Liste_Victimes)))
        Fiche_Victime = GUI_Victime_Edition(Liste_Victimes[len(Liste_Victimes)-1])
        DialogVictime.SetTopWindow(Fiche_Victime)
        Fiche_Victime.Show()
        DialogVictime.MainLoop()
        event.Skip()
    def DoubleClic(self, event):  # wxGlade: MainCourante_Fenetre.<event_handler>
        logger.debug("Event handler 'DoubleClic'")
        ligne = event.GetItem().GetId()
        logger.debug("Selection : " + str(ligne))
        gettext.install("DialogVictime")
        DialogVictime = wx.App()
        Fiche_Victime = GUI_Victime_Edition(Liste_Victimes[ligne])
        DialogVictime.SetTopWindow(Fiche_Victime)
        Fiche_Victime.Show()
        DialogVictime.MainLoop()
        event.Skip()
    def Raffraichir(self, event):  # wxGlade: MainCourante_Fenetre.<event_handler>
        logger.debug("Event handler 'Raffraichir' ")
        if self.list_ctrl_ListeVictime.GetItemCount() < len(Liste_Victimes):
            logger.debug("Affichage de la victime enregistrée à l'instant")
            self.Insere_Victime(Liste_Victimes[len(Liste_Victimes)-1], len(Liste_Victimes))
        else:
            for i in range(len(Liste_Victimes)):
                self.MetAJour_Victime( Liste_Victimes[i], i)
        event.Skip()
    def Ouvrir(self, event):  # wxGlade: MainCourante_Fenetre.<event_handler>
        logger.debug("Event handler 'Ouvrir'!")
        wb = openpyxl.load_workbook('MainCourante.xlsx', guess_types = False, data_only=True)
        ws = wb.get_sheet_by_name('MainCourante')
        i = 1
        while ws.cell(row=i, column=1).value != None:
            logger.debug(ws.cell(row=i, column=1).value)
            i += 1
        i -= 2
        logger.debug("Nombre de victime dans le fichier xlsx : " + str(i))
        Liste_Victimes = []
        for k in range(i-1):
            Liste_Victimes.append(Victime())
            Liste_Victimes[k].Dossier_num     = unicode(ws.cell(row=k+2, column=1).value)
            Liste_Victimes[k].Arrivee_HH      = unicode(ws.cell(row=k+2, column=2).value)
            Liste_Victimes[k].Arrivee_MM      = unicode(ws.cell(row=k+2, column=3).value)
            Liste_Victimes[k].Origine         = unicode(ws.cell(row=k+2, column=4).value)
            Liste_Victimes[k].Destination     = unicode(ws.cell(row=k+2, column=5).value)
            Liste_Victimes[k].Nom             = unicode(ws.cell(row=k+2, column=6).value)
            Liste_Victimes[k].Prenom          = unicode(ws.cell(row=k+2, column=7).value)
            Liste_Victimes[k].Circonstances   = unicode(ws.cell(row=k+2, column=8).value)
            Liste_Victimes[k].Traitement      = unicode(ws.cell(row=k+2, column=9).value)
            Liste_Victimes[k].Numero_Bilan    = unicode(ws.cell(row=k+2, column=10).value)
            Liste_Victimes[k].Numero_Decharge = unicode(ws.cell(row=k+2, column=11).value)
            Liste_Victimes[k].Depart_HH       = unicode(ws.cell(row=k+2, column=12).value)
            Liste_Victimes[k].Depart_MM       = unicode(ws.cell(row=k+2, column=13).value)
        for i in range(len(Liste_Victimes)):
            self.MetAJour_Victime( Liste_Victimes[i], i)

        logger.debug("fin a i = " + str(i))
        event.Skip()
    def Enregistrer(self, event):  # wxGlade: MainCourante_Fenetre.<event_handler>
        logger.debug("Event handler 'Enregistrer'!")
        wb = openpyxl.Workbook()
        ws = wb.create_sheet("MainCourante",0)
        ws.cell(row=1, column=1).value = "Dossier"
        ws.column_dimensions["A"].width = 7
        ws.cell(row=1, column=2).value = "Arr. HH"
        ws.column_dimensions["B"].width = 7
        ws.cell(row=1, column=3).value = "Arr. MM"
        ws.column_dimensions["C"].width = 7
        ws.cell(row=1, column=3).value = "ORIGINE"
        ws.column_dimensions["D"].width = 20
        ws.cell(row=1, column=5).value = "DESTINATION"
        ws.column_dimensions["E"].width = 20
        ws.cell(row=1, column=6).value = "NOM"
        ws.column_dimensions["F"].width = 15
        ws.cell(row=1, column=7).value = "PRENOM"
        ws.column_dimensions["G"].width = 15
        ws.cell(row=1, column=8).value = "CIRCONSTANCES"
        ws.column_dimensions["H"].width = 30
        ws.cell(row=1, column=9).value = "TRAITEMENT"
        ws.column_dimensions["I"].width = 100
        ws.cell(row=1, column=10).value = "BILAN#"
        ws.column_dimensions["J"].width = 12
        ws.cell(row=1, column=11).value = "DECHARGE#"
        ws.column_dimensions["K"].width = 12
        ws.cell(row=1, column=12).value = "Dép HH"
        ws.column_dimensions["K"].width = 7
        ws.cell(row=1, column=13).value = "Dép MM"
        ws.column_dimensions["M"].width = 7
        for i in range(len(Liste_Victimes)):
            ws.cell(row=i+2, column=1).value = unicode(Liste_Victimes[i].Dossier_num)
            ws.cell(row=i+2, column=2).value = unicode(Liste_Victimes[i].Arrivee_HH)
            ws.cell(row=i+2, column=3).value = unicode(Liste_Victimes[i].Arrivee_MM)
            ws.cell(row=i+2, column=4).value = unicode(Liste_Victimes[i].Origine)
            ws.cell(row=i+2, column=5).value = unicode(Liste_Victimes[i].Destination)
            ws.cell(row=i+2, column=6).value = unicode(Liste_Victimes[i].Nom)
            ws.cell(row=i+2, column=7).value = unicode(Liste_Victimes[i].Prenom)
            ws.cell(row=i+2, column=8).value = unicode(Liste_Victimes[i].Circonstances)
            ws.cell(row=i+2, column=9).value = unicode(Liste_Victimes[i].Traitement)
            ws.cell(row=i+2, column=10).value = unicode(Liste_Victimes[i].Numero_Bilan)
            ws.cell(row=i+2, column=11).value = unicode(Liste_Victimes[i].Numero_Decharge)
            ws.cell(row=i+2, column=12).value = unicode(Liste_Victimes[i].Depart_HH)
            ws.cell(row=i+2, column=13).value = unicode(Liste_Victimes[i].Depart_MM)
        wb.save('MainCourante.xlsx')
        wb.save('MainCourante.crf65.dps')
        event.Skip()


def ajout_logger():
    """ Ajoute un logger
    Un formatteur ecrit sur la console
    Un formatteur ecrit dans activity.log
    Un formatteur pourrait envoyer des emails
    http://sametmax.com/ecrire-des-logs-en-python/
    """
    _niveau_de_logging = logging.DEBUG
    _logger = logging.getLogger()
    _logger.setLevel(_niveau_de_logging)

    formatter = \
        logging.Formatter('%(asctime)s :: %(levelname)s :: %(message)s')
    file_handler = RotatingFileHandler('activity.log', 'a', 1000000, 1)
    file_handler.setLevel(logging.DEBUG)
    file_handler.setFormatter(formatter)

    _logger.addHandler(file_handler)
    steam_handler = logging.StreamHandler()
    steam_handler.setLevel(logging.INFO)
    _logger.addHandler(steam_handler)
    _logger.info("\nSTART logger au niveau : " + str(_niveau_de_logging) + "  ; (10=DEBUG; 20=INFO; 30=WARNING)")
    return _logger
def main_test_GUI_Victime_Edition():
    gettext.install("DialogVictime")
    victime_1 = Victime()
    victime_1.Set_paul_durant("1")

    DialogVictime = wx.App()
    Fiche_Victime = GUI_Victime_Edition(victime_1)
    DialogVictime.SetTopWindow(Fiche_Victime)
    Fiche_Victime.Show()
    DialogVictime.MainLoop()
def main_TEST_MainCourante_Fenetre():
    gettext.install("MainCourante")
    MainCourante = wx.App()
    frame_1 = MainCourante_Fenetre(None, wx.ID_ANY, "")
    MainCourante.SetTopWindow(frame_1)
    frame_1.Show()
    MainCourante.MainLoop()
def main_TEST_GUI_MainCourante_Fenetre():
    gettext.install("MainCourante_GUI")
    MainCourante_App     = wx.App()
    MainCourante_App_GUI = GUI_MainCourante_Fenetre()
    MainCourante_App.SetTopWindow(MainCourante_App_GUI)
    Liste_Victimes.append(Victime())
    Liste_Victimes[0].Set_defaut("1")
    Liste_Victimes.append(Victime())
    Liste_Victimes[1].Set_paul_durant("2")
    MainCourante_App_GUI.Insere_Victime(Liste_Victimes[0], 0)
    MainCourante_App_GUI.Insere_Victime(Liste_Victimes[1], 1)
    MainCourante_App_GUI.Show()
    MainCourante_App.MainLoop()




if __name__ == "__main__":
    logger = ajout_logger()
    Liste_Victimes = []
    logger.debug("Programme MainCouranteCRF65.py lancé")
#   main_test_GUI_Victime_Edition()
    main_TEST_GUI_MainCourante_Fenetre()
